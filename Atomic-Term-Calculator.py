import math
from itertools import combinations
import numpy as np
import copy
import openpyxl
from openpyxl.styles import Alignment, PatternFill, Font, Border, Side
import pandas as pd
from pyfiglet import Figlet
from time import sleep


def create_big_text(text, font='standard'):
    fig = Figlet(font=font)
    big_text = fig.renderText(text)
    return big_text
text_to_display = 'Atomic Term Calculator'
big_text = create_big_text(text_to_display, font='digital')
print(big_text)
print('This program recieves an electron configuration and gives you:')
print('| 1) Atomic Terms')
print('| 2) Microstates Table')
print('to you in two individual excel files!\n')
print('Please type the valence electron configuration:')
print('The configuration should be in this format --> ex: d2 or p3')
raw_config = input('==>\t')



def config_obtain():
    global raw_config
    for char in raw_config:
        if char.isdigit():
            q = char
        else:
            orbital = char

    return orbital, int(q)

orbital, q = config_obtain()

# Obtain z and possible ml and ms
one_electron_states = []
ms = [0.5, -0.5]
if orbital == 's':
    z = 1
    ml = [0]
elif orbital == 'p':
    z = 3
    ml = [-1, 0, 1]
elif orbital == 'd':
    z = 5
    ml = [-2, -1, 0, 1, 2]
elif orbital == 'f':
    z = 7
    ml = [-3, -2, -1, 0, 1, 2, 3]

# Obtain the lower q
if q > z:
    q = 2*z - q
else:
    pass

# Calculate the microstates
def calculate_microstates(one_electron_states, q):
    # Using nested loops to iterate through every combination
    for item1 in ml:
        for item2 in ms:
            one_electron_states.append([item1, item2])
    # Choose q-membered tuples
    unique_combinations = list(combinations(one_electron_states, q))
    return unique_combinations

unique_combinations = calculate_microstates(one_electron_states, q)

def calculate_ML_MS(unique_combinations, q):
    ML_list = []
    MS_list = []
    # Calculate ML & MS
    for microstate in unique_combinations:
        ML = 0
        MS = 0
        for i in range(q):
            temp_ml = microstate[i][0]
            temp_ms = microstate[i][1]
            ML += temp_ml
            MS += temp_ms
        ML_list.append(ML)
        MS_list.append(MS)

    return ML_list, MS_list


def calculate_atomic_term(ML_list, unique_combinations):
    L = max(ML_list)
    MS_list = []
    temp_ms = 0
    for microstate in unique_combinations:
        ML = 0
        MS = 0
        for i in range(q):
            temp_ml = microstate[i][0]
            ML += temp_ml
        if ML == L:
            for i in range(q):
                temp_ms = microstate[i][1]
                MS += temp_ms
            MS_list.append(MS)
    S = max(MS_list)
    Spin_Multiplicity = int(2*S + 1)
    terms_dictionary = {'S': 0, 'P': 1, 'D': 2, 'F': 3, 'G': 4, 'H': 5, 'I':6, 'J':7, 'K':8, 'L':9, 'M':10, 'N':11}
    # Reverse lookup to find the key (string) corresponding to the value (integer)0
    term_string = next(
        key for key, value in terms_dictionary.items() if value == L)
    atomic_term_microstate_count = int((2*L + 1)*(2*S + 1))
    return term_string, Spin_Multiplicity, atomic_term_microstate_count


def term_microstate_count(lastTerm, Spin_Multiplicity):
    terms_dictionary = {'S': 0, 'P': 1, 'D': 2, 'F': 3, 'G': 4, 'H': 5, 'I':6, 'J':7, 'K':8, 'L':9, 'M':10, 'N':11}
    L = terms_dictionary.get(lastTerm)
    S = (Spin_Multiplicity - 1)/2
    term_microstates = int((2*L + 1)*(2*S + 1))
    return term_microstates

def generate_list(L, S):
    result_list = []
    current_value = L + S

    while current_value >= L - S:
        result_list.append(current_value)
        current_value -= 1

    return result_list

def J_states_calculate(lastTerm, Spin_Multiplicity):
    terms_dictionary = {'S': 0, 'P': 1, 'D': 2, 'F': 3, 'G': 4, 'H': 5, 'I':6, 'J':7, 'K':8, 'L':9, 'M':10, 'N':11}
    L = terms_dictionary.get(lastTerm)
    S = (Spin_Multiplicity - 1)/2
    
    return generate_list(L,S)

def identify_number_type(number):
    if isinstance(number, int) or (isinstance(number, float) and number.is_integer()):
        return "Integer"
    elif isinstance(number, float):
        return "Float"
        
def remove_repeated_microstates(lastTerm, Spin_Multiplicity):
    global unique_combinations
    new_combinations = copy.copy(unique_combinations)
    terms_dictionary = {'S': 0, 'P': 1, 'D': 2, 'F': 3, 'G': 4, 'H': 5, 'I':6, 'J':7, 'K':8, 'L':9, 'M':10, 'N':11}
    L = terms_dictionary.get(lastTerm)
    S = (Spin_Multiplicity - 1)/2
    PossibleMLs = list(range(-L, L + 1))

    if identify_number_type(S) == "Float":
        PossibleMSs = list(np.arange(-S, S+0.5, 1))

    else:
        S = int(S)
        PossibleMSs = list(range(-S, S + 1))

    MS_ML_pair = [[a, b] for a in PossibleMLs for b in PossibleMSs]

    for microstate in unique_combinations:
        ML = 0
        MS = 0
        temp_ml = 0
        temp_ms = 0
        for i in range(q):
            temp_ml = microstate[i][0]
            ML += temp_ml
        for i in range(q):
            temp_ms = microstate[i][1]
            MS += temp_ms

        if [ML, MS] in MS_ML_pair:
            new_combinations.remove(microstate)
            # print(len(unique_combinations),' is what we have left')
            MS_ML_pair.remove([ML, MS])
        else:
            pass

    unique_combinations = new_combinations


# Find the total number of states
def state_calculator(z, q):
    n = (math.factorial(2*z)/(math.factorial(q)*math.factorial(2*z-q)))
    return int(n)


def microstate_ML_MS_finder(ML_list, MS_list):
    microstate_ML_MS_list = []
    z = len(ML_list)
    for i in range(z):
        microstate_ML_MS_list.append([ML_list[i],MS_list[i]])
    return microstate_ML_MS_list


# Count microstates with specific ML&MS
def specific_microstate_counter(ML, MS, microstate_ML_MS_list):
    count = 0
    for microstate in microstate_ML_MS_list:
        # The same
        if microstate[0] == ML and microstate[1] == MS:
            count += 1
        else:
            pass
    return count

def microstates_to_excel(microstate_ML_MS_list, ML_list, MS_list):

    # Obtain headers
    unique_ML_list = []
    for ML in ML_list:
        if ML not in unique_ML_list:
            unique_ML_list.append(ML)

    unique_MS_list = []
    for MS in MS_list:
        if MS not in unique_MS_list:
            unique_MS_list.append(MS)

    unique_ML_list = sorted(unique_ML_list)
    unique_MS_list = sorted(unique_MS_list)

    # Count microstates with specific ML&MS
    # Create a new workbook
    workbook = openpyxl.Workbook()

    # Create a new sheet
    sheet = workbook.active
    sheet.title = "MicroStates"

    # Define header names
    header_names = ["ML\\MS"] +  [MS for MS in unique_MS_list]
    # Populate the first row with header names
    for col, header in enumerate(header_names, start=1):
        sheet.cell(row=1, column=col, value=header)

    for row in range(2, len(unique_ML_list)+2):
        # Add row name in the first column
        row_name = unique_ML_list[row-2]
        sheet.cell(row=row, column=1, value=row_name)
        
        for col in range(2, len(unique_MS_list)+2):
            cell_value = specific_microstate_counter(ML=unique_ML_list[row-2], MS=unique_MS_list[col-2], microstate_ML_MS_list=microstate_ML_MS_list)
            # print(cell_value)
            sheet.cell(row=row, column=col, value=cell_value)

    
    # Set alignment to center for all cells
    for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row, min_col=1, max_col=sheet.max_column):
        for cell in row:
            cell.alignment = Alignment(horizontal='center', vertical='center')

    # Set background color, bold text, and white borders for the first row (headers) and the first column
    for row in sheet.iter_rows(min_row=1, max_row=1, min_col=1, max_col=sheet.max_column):
        for cell in row:
            cell.fill = PatternFill(start_color='000000', end_color='000000', fill_type='solid')
            cell.font = Font(color='FFFFFF', bold=True)
            cell.border = Border(left=Side(style='thin', color='FFFFFF'), 
                                right=Side(style='thin', color='FFFFFF'), 
                                top=Side(style='thin', color='FFFFFF'), 
                                bottom=Side(style='thin', color='FFFFFF'))

    for col in sheet.iter_cols(min_row=2, max_row=sheet.max_row, min_col=1, max_col=1):
        for cell in col:
            cell.fill = PatternFill(start_color='000000', end_color='000000', fill_type='solid')
            cell.font = Font(color='FFFFFF', bold=True)
            cell.border = Border(left=Side(style='thin', color='FFFFFF'), 
                                right=Side(style='thin', color='FFFFFF'), 
                                top=Side(style='thin', color='FFFFFF'), 
                                bottom=Side(style='thin', color='FFFFFF'))

    # Set minimum width for each column
    for col in sheet.columns:
        max_length = 0
        column = [cell for cell in col]
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2)
        sheet.column_dimensions[col[0].column_letter].width = adjusted_width if adjusted_width > 10 else 10


    # Save the workbook
    workbook.save('Microstates Table.xlsx')



atomic_terms_list = []
atomic_term_count_list = []
all_J_list = []
x = 1
mircorstates_table_drawn = False

while x != 0:
    ML_list, MS_list = calculate_ML_MS(unique_combinations, q)
    raw_MS = sorted(list(set(MS_list)))
    raw_ML = sorted(list(set(ML_list)))
    if not mircorstates_table_drawn:
        microstate_ML_MS_list = microstate_ML_MS_finder(ML_list, MS_list)
        microstates_to_excel(microstate_ML_MS_list, ML_list, MS_list)
        mircorstates_table_drawn = True
    else:
        pass  
    term_symbol, Spin_Multiplicity, atomic_term_microstate_count = calculate_atomic_term(
        ML_list, unique_combinations)
    atomic_terms_list.append([Spin_Multiplicity, term_symbol])
    atomic_term_count_list.append(atomic_term_microstate_count)
    term_microstates = term_microstate_count(term_symbol, Spin_Multiplicity)
    J_states_list = J_states_calculate(term_symbol, Spin_Multiplicity)
    all_J_list.append(J_states_list)
    remove_repeated_microstates(term_symbol, Spin_Multiplicity)
    x = len(unique_combinations)


new_term_list = []
for term in atomic_terms_list:
    full_term = f'{term[0]}{term[1]}'
    new_term_list.append(full_term)
terms_excel = {'Atomic Terms': [term for term in new_term_list],
               'Substates (J)': [str(Jlist) for Jlist in all_J_list],
               'Microstate Count': [count for count in atomic_term_count_list]}
terms_excel['Total Microstates'] = [state_calculator(
    z, q)] + [''] * (len(terms_excel['Atomic Terms']) - 1)
terms_df = pd.DataFrame(terms_excel)
column_width = 100
terms_excel_path = 'Atomic Terms.xlsx'


with pd.ExcelWriter(terms_excel_path, engine='xlsxwriter') as writer:
    terms_df.to_excel(writer, index=False, sheet_name='Sheet1')
    workbook = writer.book
    worksheet = writer.sheets['Sheet1']
    header_format = workbook.add_format(
        {'bold': True, 'bg_color': 'darkred', 'text_wrap': True, 'align': 'center', 'valign': 'vcenter', 'color': 'white'})
    for col_num, value in enumerate(terms_df.columns.values):
        worksheet.write(0, col_num, value, header_format)
    for i, col in enumerate(terms_df.columns):
        column_length = max(terms_df[col].astype(
            str).apply(len).max(), len(col) + 2)
        worksheet.set_column(i, i, min(column_length, column_width))
    data_format = workbook.add_format({'align': 'center', 'valign': 'vcenter'})
    for row_num, values in enumerate(terms_df.values, 1):
        for col_num, value in enumerate(values):
            worksheet.write(row_num, col_num, value, data_format)
sleep(0.5)
print('\n----> Excels Created! <----')
print('\n========= Created by Milad =========')
print('Github:     @chemistmilad')
print('Telegram:   @imjustmilad')
input()